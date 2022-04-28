from ctypes import create_unicode_buffer
from tkinter import mainloop
import tkinter
from tokenize import String
from typing import List
from selenium import webdriver
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from re import search
import time
import os
import root
import pandas as pd
from pathlib import Path
import openpyxl
import win32com.client

#Variável com nome da pasta que será criada


#Verificação se a pasta existe
def criarPasta() -> str:
    dir = 'C:/CSV'
    for dic in os.listdir('C:/'):
        if os.path.isdir(dir):
            for txt in os.listdir(dir):
                os.remove(os.path.join(dir, txt))
            print(f'Arquivos da pasta {dir} excluidos')
            break
        else:
            os.mkdir(dir)
            print(f'Pasta {dir} criada')
    return dir

def ajustaData() -> str:
    app = root.Janela()
    tkinter.mainloop()
    data = app.getInput()
    data.replace("/","")
    return data

#Confirguração Firefox
def configuraFirefox() -> webdriver:
    fp = webdriver.FirefoxProfile()
    fp.set_preference("browser.download.dir", os.path.join('c:','CSV'))
    fp.set_preference("browser.preferences.instantApply", True)
    fp.set_preference("browser.helperApps.neverAsk.saveToDisk",
                    "text/plain, application/octet-stream, application/binary, text/csv, application/csv, application/excel, text/comma-separated-values, text/xml, application/xml")
    fp.set_preference("browser.helperApps.alwaysAsk.force", False)
    fp.set_preference("browser.download.manager.showWhenStarting", False)
    fp.set_preference("browser.download.folderList", 2)
    return fp

#Navegação
def navegaSite() -> None:
    fp = configuraFirefox()
    data = ajustaData()
    navegador = webdriver.Firefox(firefox_profile=fp)
    navegador.get('www.site.com.br')
    username = navegador.find_element_by_css_selector('#txtmatricula')
    username.clear()
    username.send_keys("login")
    password = navegador.find_element_by_css_selector('#txtsenha')
    password.clear()
    password.send_keys("senha")
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#btnLogin'))).click()
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'li.static:nth-child(2) > a:nth-child(1)'))).click()
    time.sleep(2)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.select2-selection__clear'))).click()
    time.sleep(0.5)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#select2-MainContent_cbTipoEmpresa-container'))).click()
    navegador.find_element_by_css_selector('.select2-search__field').send_keys(Keys.ENTER)
    time.sleep(1)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#select2-MainContent_cbempresa-container'))).click()
    navegador.find_element_by_css_selector('.select2-search__field').send_keys(Keys.ENTER)
    time.sleep(1)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#MainContent_chkencerrada'))).click()
    time.sleep(1)
    txtdata = navegador.find_element_by_css_selector("#MainContent_txtDataInicial")
    txtdata.send_keys(data)
    txtdata = navegador.find_element_by_css_selector("#MainContent_txtDataFinal")
    txtdata.send_keys(data)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#MainContent_ButtonPesquisar'))).click()
    time.sleep(10)
    WebDriverWait(navegador, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#MainContent_btnExportar'))).click()
    time.sleep(15)
    navegador.close()

def selecionaArquivo() -> list:
    home = str(Path.home())
    search_dir = os.path.join(home,'Downloads')
    files = list(os.listdir(search_dir))
    files = list(filter(lambda file: file.endswith('.csv'), files))
    files = [os.path.join(search_dir, f) for f in files]
    files.sort(key = os.path.getmtime)
    return home, search_dir, files

#Manipulando excel
def excel() -> None:
    lista_arquivos = selecionaArquivo()
    files = lista_arquivos[2]
    search_dir = lista_arquivos[1]
    df = pd.read_csv(files[-1], sep=';',decimal=',')
    df.drop(df.columns[[1,4,10,11,12,13]],axis=1,inplace=True)
    df        = df.astype({'CUSTO_MEDIO_DIVERGENTE': 'float64'})
    df_taqi   = df[df['NOME_UNIDADE_NEGOCIO'].astype(str).str.contains('TAQI')]
    df_iplace = df[df['NOME_UNIDADE_NEGOCIO'].astype(str).str.contains('IPLACE')]
    df_5      = df_taqi[df_taqi['CUSTO_MEDIO_DIVERGENTE']<-5000]
    df_20     = df_iplace[df_iplace['CUSTO_MEDIO_DIVERGENTE']<-20000]

    df_5_x  = df_5['NOME_UNIDADE_NEGOCIO'].str.split(' - ', expand=True)
    df_20_x = df_20['NOME_UNIDADE_NEGOCIO'].str.split(' - ', expand=True)

    df_5  = pd.concat([df_5,df_5_x], axis=1)
    df_20 = pd.concat([df_20,df_20_x], axis=1)

    df_5.to_excel(os.path.join(search_dir,'filiais_taqi.xlsx'),index=False)
    df_20.to_excel(os.path.join(search_dir,'filiais_iplace.xlsx'),index=False)

    wb    = openpyxl.load_workbook('Envio de e-mails.xlsm',read_only=False, keep_vba=True)
    ws    = wb.worksheets[0]
    ws["A1"] = str(os.path.join(search_dir,'filiais_taqi.xlsx'))
    ws["A2"] = str(os.path.join(search_dir,'filiais_iplace.xlsx'))
    wb.save('Envio de e-mails.xlsm')
    wb.close()

    if os.path.exists("Envio de e-mails.xlsm"): 
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(os.path.abspath("Envio de e-mails.xlsm"))
        xl.Visible = True
        try:
            xl.Application.Run("'Envio de e-mails.xlsm'!importar_dados")
        except:
            xl.Application.Quit()
        del xl